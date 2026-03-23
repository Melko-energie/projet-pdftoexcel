"""Création de fichiers Excel formatés professionnellement."""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .parser import Dataset

logger = logging.getLogger(__name__)

# Styles
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROW_EVEN_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ROW_ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)

CELL_FONT = Font(name="Calibri", size=11)
CELL_ALIGNMENT = Alignment(vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

EURO_FORMAT = '#,##0.00 €'
DATE_FORMAT = 'DD/MM/YYYY'


def _sanitize_sheet_name(name: str) -> str:
    """Nettoie un nom de feuille Excel (max 31 caractères, pas de caractères interdits)."""
    forbidden = r"[]:*?/\\"
    clean = "".join(c for c in name if c not in forbidden)
    return clean[:31]


def _auto_column_width(ws, col_idx: int, header: str, values: list) -> float:
    """Calcule la largeur optimale d'une colonne."""
    max_len = len(str(header))
    for val in values[:100]:  # Limiter le scan aux 100 premières lignes
        if val is not None:
            cell_len = len(str(val))
            if cell_len > max_len:
                max_len = cell_len
    # Ajouter un peu de marge
    return min(max_len + 4, 50)


def write_dataset_to_sheet(ws, dataset: Dataset) -> None:
    """Écrit un Dataset dans une feuille Excel avec formatage."""
    headers = dataset.headers
    col_types = dataset.column_types
    num_cols = len(headers)

    # Écriture des headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Écriture des données
    current_row = 2
    for row_idx, row in enumerate(dataset.data_rows):
        fill = ROW_EVEN_FILL if row_idx % 2 == 0 else ROW_ODD_FILL
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.fill = fill
            cell.border = THIN_BORDER

            # Appliquer le format selon le type de colonne
            if col_idx <= len(col_types):
                col_type = col_types[col_idx - 1]
                if col_type == "euro" and isinstance(value, (int, float)):
                    cell.number_format = EURO_FORMAT
                elif col_type == "date" and isinstance(value, datetime):
                    cell.number_format = DATE_FORMAT
        current_row += 1

    # Écriture des lignes de total (en bas, avec style distinct)
    for total_row in dataset.total_rows:
        for col_idx, value in enumerate(total_row, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

            if col_idx <= len(col_types):
                col_type = col_types[col_idx - 1]
                if col_type == "euro" and isinstance(value, (int, float)):
                    cell.number_format = EURO_FORMAT
                elif col_type == "date" and isinstance(value, datetime):
                    cell.number_format = DATE_FORMAT
        current_row += 1

    # Auto-ajustement de la largeur des colonnes
    for col_idx in range(1, num_cols + 1):
        header = headers[col_idx - 1]
        col_values = [
            row[col_idx - 1]
            for row in dataset.data_rows
            if col_idx - 1 < len(row)
        ]
        width = _auto_column_width(ws, col_idx, header, col_values)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes sur la première ligne
    ws.freeze_panes = "A2"

    # Auto-filtre
    if num_cols > 0 and current_row > 1:
        last_col = get_column_letter(num_cols)
        ws.auto_filter.ref = f"A1:{last_col}{current_row - 1}"


def write_excel(datasets: list[Dataset], output_path: Path) -> Path:
    """Crée un fichier Excel à partir des Datasets.

    Args:
        datasets: Liste de Datasets à écrire.
        output_path: Chemin du fichier de sortie.

    Returns:
        Le chemin du fichier créé.
    """
    if not datasets:
        logger.warning("Aucun dataset à écrire.")
        return output_path

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    for dataset in datasets:
        sheet_name = _sanitize_sheet_name(dataset.name)
        ws = wb.create_sheet(title=sheet_name)

        logger.info(
            "Écriture de la feuille '%s' : %d lignes + %d totaux",
            sheet_name, len(dataset.data_rows), len(dataset.total_rows),
        )

        write_dataset_to_sheet(ws, dataset)

    wb.save(output_path)
    logger.info("Fichier Excel créé : %s", output_path)
    return output_path
